from __future__ import annotations

import io
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from hr_analytics.excel_report import (
    TURNOVER_COST_MULTIPLIER_BY_LEVEL,
    estimate_turnover_cost,
    generate_report,
)

EXPECTED_SHEETS = [
    "Executive Summary",
    "Attrition by Dept & Role",
    "Department x Tenure Pivot",
    "Flight Risk Watchlist",
    "Survival Model Hazard Ratios",
    "Hiring Pipeline (Synthetic)",
]


def test_estimate_turnover_cost_known_fixture() -> None:
    fixture = pd.DataFrame(
        [
            {"EmployeeNumber": 1, "Department": "Sales", "JobRole": "Rep", "JobLevel": 2, "MonthlyIncome": 5000, "Attrition": "Yes"},
            {"EmployeeNumber": 2, "Department": "Sales", "JobRole": "Rep", "JobLevel": 5, "MonthlyIncome": 10000, "Attrition": "Yes"},
            {"EmployeeNumber": 3, "Department": "Sales", "JobRole": "Rep", "JobLevel": 1, "MonthlyIncome": 3000, "Attrition": "No"},
        ]
    )
    result = estimate_turnover_cost(fixture)

    assert len(result) == 2  # only the two leavers, the stayer is excluded
    assert set(result["EmployeeNumber"]) == {1, 2}

    row1 = result.loc[result["EmployeeNumber"] == 1].iloc[0]
    expected_1 = 5000 * 12 * TURNOVER_COST_MULTIPLIER_BY_LEVEL[2]
    assert row1["estimated_turnover_cost"] == pytest.approx(expected_1)

    row2 = result.loc[result["EmployeeNumber"] == 2].iloc[0]
    expected_2 = 10000 * 12 * TURNOVER_COST_MULTIPLIER_BY_LEVEL[5]
    assert row2["estimated_turnover_cost"] == pytest.approx(expected_2)
    # senior hire's estimated cost should exceed the junior hire's despite
    # a lower income gap than the multiplier gap -- confirms the multiplier
    # is actually being applied, not just annual salary alone
    assert row2["estimated_turnover_cost"] > row1["estimated_turnover_cost"]


def test_estimate_turnover_cost_empty_when_no_leavers() -> None:
    fixture = pd.DataFrame(
        [{"EmployeeNumber": 1, "Department": "Sales", "JobRole": "Rep", "JobLevel": 2, "MonthlyIncome": 5000, "Attrition": "No"}]
    )
    result = estimate_turnover_cost(fixture)
    assert len(result) == 0


@pytest.fixture(scope="module")
def report_bytes() -> bytes:
    return generate_report(output_path=None)


def test_generate_report_produces_valid_workbook(report_bytes: bytes) -> None:
    assert len(report_bytes) > 10_000  # a near-empty/corrupt file would be far smaller

    wb = openpyxl.load_workbook(io.BytesIO(report_bytes))
    assert wb.sheetnames == EXPECTED_SHEETS


def test_generate_report_does_not_write_to_disk_when_output_path_is_none(tmp_path: Path) -> None:
    import hr_analytics.excel_report as excel_report

    before = set(tmp_path.iterdir())
    excel_report.generate_report(output_path=None)
    after = set(tmp_path.iterdir())
    assert before == after


def test_generate_report_writes_to_disk_when_output_path_given(tmp_path: Path) -> None:
    out = tmp_path / "report.xlsx"
    generate_report(output_path=out)
    assert out.exists()
    assert out.read_bytes() == generate_report(output_path=None)


def test_executive_summary_matches_known_attrition_rate(report_bytes: bytes) -> None:

    wb = openpyxl.load_workbook(io.BytesIO(report_bytes))
    ws = wb["Executive Summary"]
    # matches docs/sql_findings.md's documented overall attrition rate
    assert ws["B5"].value == "16.1%"
    assert ws["B3"].value == 1470  # headcount


def test_conditional_formatting_actually_attached(report_bytes: bytes) -> None:

    wb = openpyxl.load_workbook(io.BytesIO(report_bytes))
    for sheet_name in ["Attrition by Dept & Role", "Flight Risk Watchlist", "Survival Model Hazard Ratios"]:
        ws = wb[sheet_name]
        assert len(list(ws.conditional_formatting)) > 0, f"{sheet_name} has no conditional formatting attached"


def test_report_generation_is_byte_reproducible() -> None:
    """Regression test for two real bugs found while building this module:
    openpyxl.writer.excel.save_workbook() unconditionally stamps
    docProps/core.xml's <dcterms:modified> with datetime.now() at save
    time (worked around by post-processing the zip), and the zip
    container's own per-entry DOS timestamps are a second, independent
    non-determinism source. Confirms both fixes hold, not just that the
    file is valid."""
    first = generate_report(output_path=None)
    second = generate_report(output_path=None)
    assert first == second
