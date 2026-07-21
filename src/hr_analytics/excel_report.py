"""Generates a formatted Excel executive report from this project's existing
pipeline outputs -- the kind of stakeholder-ready deliverable an HR/People
Ops leader would actually get handed, distinct from the live dashboard.

Six sheets: Executive Summary (KPI scorecard, including an estimated
turnover cost -- see estimate_turnover_cost()'s docstring and DECISIONS.md
for the methodology and its caveats), Attrition by Department & Role,
a Department x Tenure pivot-style summary, Flight Risk Watchlist, Survival
Model Hazard Ratios, and Hiring Pipeline (synthetic, clearly labelled).

"Pivot table" here means a pandas-computed pivot_table() written as a
static formatted table, not a genuine interactive native Excel PivotTable
-- neither openpyxl nor xlsxwriter can reliably create those from scratch.
See DECISIONS.md for why that's the honest scope, not an oversight.

generate_report() always returns the workbook's bytes and optionally also
writes them to disk, so dashboard/app.py's download button and this
module's own CLI entry point share one code path instead of building the
workbook twice.
"""

from __future__ import annotations

import io
import json
import re
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ENRICHED_PATH = Path("data/processed/hr_employee_attrition_enriched.csv")
RISK_PATH = Path("data/processed/predicted_attrition_risk.csv")
COEFFICIENTS_PATH = Path("data/processed/survival_model_coefficients.csv")
METRICS_PATH = Path("data/processed/survival_model_metrics.json")
ATTRITION_BY_DEPT_ROLE_PATH = Path("sql/results/02_attrition_by_department_and_role.csv")
HIRING_BY_LEVEL_PATH = Path("sql/results/08_time_to_hire_by_department_and_level.csv")
OUTPUT_PATH = Path("reports/HR_Executive_Report.xlsx")

# Fixed, not datetime.now(): openpyxl stamps the workbook's core properties
# with the current wall-clock time by default, which would make every
# regeneration byte-different even with identical data. See DECISIONS.md.
#
# Setting wb.properties.created/.modified before save() only half-works:
# openpyxl.writer.excel.save_workbook() unconditionally overwrites
# `workbook.properties.modified = datetime.datetime.now(...)` immediately
# before writing, with no public parameter to opt out (verified by reading
# openpyxl/writer/excel.py directly, not assumed). `created` isn't touched
# there and does stay fixed. So `modified` has to be corrected by
# post-processing the already-saved zip -- see _freeze_workbook_timestamp().
_FIXED_WORKBOOK_TIMESTAMP = datetime(2026, 1, 1)
_CORE_XML_MODIFIED_PATTERN = re.compile(
    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)"
)

# Mirrors synthetic_hiring.py's BASE_MEDIAN_DAYS_BY_LEVEL pattern: a
# job-level-scaled multiplier, not a flat rate, since replacing a senior
# hire costs more (recruiting, ramp-up, lost productivity) than a junior
# one. Illustrative industry rule-of-thumb, not observed cost data for
# this (fictional) company -- see DECISIONS.md.
TURNOVER_COST_MULTIPLIER_BY_LEVEL = {1: 0.5, 2: 0.75, 3: 1.0, 4: 1.5, 5: 2.0}

TENURE_BINS = [-1, 0, 2, 4, 9, 100]
TENURE_LABELS = ["<1 year", "1-2 years", "3-4 years", "5-9 years", "10+ years"]

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_NOTE_FONT = Font(italic=True, size=9, color="595959")


def estimate_turnover_cost(employees: pd.DataFrame) -> pd.DataFrame:
    """Per-leaver estimated replacement cost = MonthlyIncome * 12 * a
    job-level-scaled multiplier, for employees where Attrition == 'Yes'
    only. Pure function over a DataFrame with EmployeeNumber, Department,
    JobRole, MonthlyIncome, JobLevel, Attrition columns -- no file I/O --
    so it's directly testable against a small known fixture."""
    leavers = employees[employees["Attrition"] == "Yes"].copy()
    multiplier = leavers["JobLevel"].map(TURNOVER_COST_MULTIPLIER_BY_LEVEL).fillna(1.0)
    leavers["estimated_annual_salary"] = leavers["MonthlyIncome"] * 12
    leavers["cost_multiplier"] = multiplier
    leavers["estimated_turnover_cost"] = leavers["estimated_annual_salary"] * multiplier
    return leavers[
        [
            "EmployeeNumber",
            "Department",
            "JobRole",
            "JobLevel",
            "estimated_annual_salary",
            "cost_multiplier",
            "estimated_turnover_cost",
        ]
    ].sort_values("estimated_turnover_cost", ascending=False)


def _style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize_columns(ws: Worksheet, df: pd.DataFrame, start_col: int = 1) -> None:
    for i, col_name in enumerate(df.columns):
        col_letter = get_column_letter(start_col + i)
        max_len = max(len(str(col_name)), df[col_name].astype(str).map(len).max() if len(df) else 0)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def _write_dataframe(
    ws: Worksheet, df: pd.DataFrame, start_row: int = 1, start_col: int = 1
) -> tuple[int, int]:
    """Writes a header row + data rows starting at (start_row, start_col).
    Returns (first_data_row, last_data_row) for the caller to apply
    conditional formatting or other range-based operations against."""
    for j, col_name in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=str(col_name))
    _style_header_row(ws, start_row, len(df.columns))

    for i, (_, row) in enumerate(df.iterrows()):
        for j, value in enumerate(row):
            if pd.isna(value):
                value = None
            elif hasattr(value, "item"):  # numpy scalar -> native Python type
                value = value.item()
            ws.cell(row=start_row + 1 + i, column=start_col + j, value=value)

    _autosize_columns(ws, df, start_col)
    return start_row + 1, start_row + len(df)


def _apply_color_scale(ws: Worksheet, col_letter: str, first_row: int, last_row: int) -> None:
    if last_row < first_row:
        return
    ws.conditional_formatting.add(
        f"{col_letter}{first_row}:{col_letter}{last_row}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        ),
    )


def load_report_data() -> dict[str, pd.DataFrame | dict]:
    return {
        "employees": pd.read_csv(ENRICHED_PATH, encoding="utf-8-sig"),
        "attrition_by_dept_role": pd.read_csv(ATTRITION_BY_DEPT_ROLE_PATH),
        "risk": pd.read_csv(RISK_PATH),
        "coefficients": pd.read_csv(COEFFICIENTS_PATH),
        "hiring_by_level": pd.read_csv(HIRING_BY_LEVEL_PATH),
        "metrics": json.loads(METRICS_PATH.read_text()),
    }


def build_executive_summary_sheet(
    wb: Workbook, employees: pd.DataFrame, metrics: dict, turnover_costs: pd.DataFrame
) -> None:
    ws = wb.create_sheet("Executive Summary")
    ws["A1"] = "IBM HR Analytics Hub -- Executive Summary"
    ws["A1"].font = _TITLE_FONT

    headcount = len(employees)
    leavers = int((employees["Attrition"] == "Yes").sum())
    attrition_rate = round(100 * leavers / headcount, 1)
    avg_tenure = round(employees["YearsAtCompany"].mean(), 1)
    total_turnover_cost = round(turnover_costs["estimated_turnover_cost"].sum())

    kpis = [
        ("Headcount", headcount),
        ("Leavers", leavers),
        ("Attrition rate", f"{attrition_rate}%"),
        ("Avg. tenure (years)", avg_tenure),
        ("Model concordance (5-fold CV)", metrics["concordance_cv_mean"]),
        ("Estimated total turnover cost", f"${total_turnover_cost:,.0f}"),
    ]
    row = 3
    for label, value in kpis:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Turnover cost is an ILLUSTRATIVE ESTIMATE").font = Font(bold=True, color="C00000")
    row += 1
    note = (
        "Computed as MonthlyIncome x 12 x a job-level-scaled multiplier "
        "(0.5x-2.0x, junior to senior), an industry rule-of-thumb -- not "
        "observed cost data for this (fictional) company. See DECISIONS.md."
    )
    ws.cell(row=row, column=1, value=note).font = _NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20


def build_attrition_by_department_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Attrition by Dept & Role")
    first_row, last_row = _write_dataframe(ws, df)
    col_letter = get_column_letter(list(df.columns).index("attrition_rate_pct") + 1)
    _apply_color_scale(ws, col_letter, first_row, last_row)


def build_tenure_department_pivot_sheet(wb: Workbook, employees: pd.DataFrame) -> None:
    ws = wb.create_sheet("Department x Tenure Pivot")
    ws["A1"] = "Attrition rate (%) by department and tenure bucket"
    ws["A1"].font = _TITLE_FONT

    tenure = employees.copy()
    tenure["tenure_bucket"] = pd.cut(tenure["YearsAtCompany"], bins=TENURE_BINS, labels=TENURE_LABELS)
    tenure["left"] = (tenure["Attrition"] == "Yes").astype(int)
    pivot = pd.pivot_table(
        tenure, index="Department", columns="tenure_bucket", values="left", aggfunc="mean", observed=True
    ) * 100
    pivot = pivot.round(1).reset_index()

    first_row, last_row = _write_dataframe(ws, pivot, start_row=3)
    for col_idx in range(2, len(pivot.columns) + 1):
        _apply_color_scale(ws, get_column_letter(col_idx), first_row, last_row)


def build_flight_risk_sheet(wb: Workbook, risk: pd.DataFrame) -> None:
    ws = wb.create_sheet("Flight Risk Watchlist")
    current = risk[risk["Attrition"] == "No"].sort_values("predicted_hazard_score", ascending=False)
    first_row, last_row = _write_dataframe(ws, current)
    col_letter = get_column_letter(list(current.columns).index("predicted_hazard_score") + 1)
    _apply_color_scale(ws, col_letter, first_row, last_row)


def build_hazard_ratios_sheet(wb: Workbook, coefficients: pd.DataFrame) -> None:
    ws = wb.create_sheet("Survival Model Hazard Ratios")
    df = coefficients[["covariate", "exp(coef)", "p"]].sort_values("exp(coef)", ascending=False)
    first_row, last_row = _write_dataframe(ws, df)
    col_letter = get_column_letter(list(df.columns).index("exp(coef)") + 1)
    if last_row >= first_row:
        ws.conditional_formatting.add(
            f"{col_letter}{first_row}:{col_letter}{last_row}",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="num", mid_value=1.0, mid_color="FFFFFF",
                end_type="max", end_color="F8696B",
            ),
        )
    note_row = last_row + 2
    ws.cell(row=note_row, column=1, value="Hazard ratio > 1 raises attrition risk; < 1 lowers it. See DECISIONS.md.").font = _NOTE_FONT


def build_hiring_pipeline_sheet(wb: Workbook, hiring_by_level: pd.DataFrame) -> None:
    ws = wb.create_sheet("Hiring Pipeline (Synthetic)")
    ws["A1"] = "SYNTHETIC DATA -- simulated, not observed hiring records. See DECISIONS.md."
    ws["A1"].font = Font(bold=True, color="C00000")
    ws.merge_cells("A1:E1")
    _write_dataframe(ws, hiring_by_level, start_row=3)


def _build_workbook(data: dict) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # default blank sheet
    wb.properties.creator = "hr-analytics-hub"
    wb.properties.created = _FIXED_WORKBOOK_TIMESTAMP
    wb.properties.modified = _FIXED_WORKBOOK_TIMESTAMP
    wb.properties.lastModifiedBy = "hr-analytics-hub"

    turnover_costs = estimate_turnover_cost(data["employees"])
    build_executive_summary_sheet(wb, data["employees"], data["metrics"], turnover_costs)
    build_attrition_by_department_sheet(wb, data["attrition_by_dept_role"])
    build_tenure_department_pivot_sheet(wb, data["employees"])
    build_flight_risk_sheet(wb, data["risk"])
    build_hazard_ratios_sheet(wb, data["coefficients"])
    build_hiring_pipeline_sheet(wb, data["hiring_by_level"])
    return wb


def _freeze_workbook_timestamp(xlsx_bytes: bytes) -> bytes:
    """Rewrites docProps/core.xml's <dcterms:modified> to
    _FIXED_WORKBOOK_TIMESTAMP inside an already-saved .xlsx, working around
    save_workbook()'s hardcoded datetime.now() (see the comment on
    _FIXED_WORKBOOK_TIMESTAMP above).

    Also resets every zip entry's own DOS-format date_time field. This is a
    second, independent non-determinism source from docProps/core.xml's XML
    content -- the .xlsx zip container itself stamps a per-entry last-modified
    timestamp at a fixed byte offset near the start of the file, which a
    content-only diff (e.g. `diff` after `unzip`) won't reveal since it's
    zip metadata, not file content. Found this by regenerating twice, seeing
    identical byte counts but a `cmp` mismatch at a low byte offset (byte 11,
    inside the very first entry's local file header) even after the XML fix
    landed -- not assumed from documentation."""
    fixed = _FIXED_WORKBOOK_TIMESTAMP.strftime("%Y-%m-%dT%H:%M:%SZ").encode()
    fixed_date_time = (
        _FIXED_WORKBOOK_TIMESTAMP.year, _FIXED_WORKBOOK_TIMESTAMP.month, _FIXED_WORKBOOK_TIMESTAMP.day,
        0, 0, 0,
    )
    src = zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r")
    out_buffer = io.BytesIO()
    with zipfile.ZipFile(out_buffer, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "docProps/core.xml":
                # \g<1>/\g<2>, not bare \1/\2: `fixed` starts with digits
                # ("2026-..."), and \1 immediately followed by digits is
                # ambiguous in a replacement string -- Python's re module
                # parsed \1 + "20" as the octal escape \120 (chr(80) = "P")
                # rather than group 1 followed by literal text, corrupting
                # the XML. Confirmed by reproducing the exact "P26-01-01..."
                # corruption in isolation before applying this fix.
                data = _CORE_XML_MODIFIED_PATTERN.sub(rb"\g<1>" + fixed + rb"\g<2>", data)
            item.date_time = fixed_date_time
            dst.writestr(item, data)
    return out_buffer.getvalue()


def generate_report(output_path: Path | None = OUTPUT_PATH) -> bytes:
    data = load_report_data()
    wb = _build_workbook(data)
    buffer = io.BytesIO()
    wb.save(buffer)
    content = _freeze_workbook_timestamp(buffer.getvalue())
    if output_path is not None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(content)
    return content


def main() -> None:
    content = generate_report()
    print(f"Wrote {len(content):,} bytes to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
